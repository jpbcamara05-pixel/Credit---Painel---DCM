#!/usr/bin/env python3
"""
DCM Credit Screening — Excel Filler v1.0
Preenche a planilha empresa por empresa com dados extraídos dos PDFs.
LTM = FY2025 (release 4T25) − 1T25 (comparativo 1T26) + 1T26 (corrente)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import shutil, os, copy

# ─── Paths ────────────────────────────────────────────────────────────────────
BASE   = "/Users/joaopedrocamara/Documents/VS/Credit Study"
INPUT  = f"{BASE}/DCM_Credit_Screening (1) (1).xlsx"
OUTPUT = f"{BASE}/output/DCM_Credit_Screening_PREENCHIDO.xlsx"

# ─── Copy template ────────────────────────────────────────────────────────────
shutil.copy2(INPUT, OUTPUT)
wb = openpyxl.load_workbook(OUTPUT)

# ─── Sheet references ─────────────────────────────────────────────────────────
ws_inputs   = wb["Inputs — DFs"]
ws_crono    = wb["Cronograma de Amortização"]
ws_wall     = wb["Wall of Maturities"]
ws_cov      = wb["Covenants & Capex"]

# ─── Create new tabs ──────────────────────────────────────────────────────────
def get_or_create_sheet(wb, name, after=None):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    return ws

ws_disc  = get_or_create_sheet(wb, "Disclaimers")
ws_deal  = get_or_create_sheet(wb, "Oportunidades de Deal")

# Style helpers
BLUE_FILL  = PatternFill("solid", fgColor="DCE6F1")
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
BOLD       = Font(bold=True)
WRAP       = Alignment(wrap_text=True, vertical="top")
CENTER     = Alignment(horizontal="center", vertical="center")

def hdr(ws, row, col, text, fill=HDR_FILL, font=HDR_FONT):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill; c.font = font; c.alignment = CENTER

def inp(ws, row, col, value):
    if value is None:
        return
    c = ws.cell(row=row, column=col, value=value)
    c.fill = BLUE_FILL

def find_row(ws, name, col=1, start=1):
    for r in ws.iter_rows(min_row=start, max_col=col, values_only=False):
        if r[col-1].value and str(r[col-1].value).strip() == name:
            return r[col-1].row
    return None

# ══════════════════════════════════════════════════════════════════════════════
#  DATA — ÂNIMA (1T26 LTM)
# ══════════════════════════════════════════════════════════════════════════════
ANIMA = {
    "name":    "Ânima",
    "sector":  "Educação",

    # ── Inputs — DFs (R$ mm) ─────────────────────────────────────────────────
    "periodo":        "1T26 LTM",
    "emp_fin_cp":      474.8,    # Emp & Fin + Debêntures CP (excl. derivativos)
    "emp_fin_lp":     4024.3,    # Emp & Fin + Debêntures LP (excl. derivativos)
    "arrendam_cp":     132.6,    # IFRS16 CP (Nota 10)
    "arrendam_lp":    1243.3,    # IFRS16 LP (Nota 10)
    "outros_pass_cp":   None,    # Sem CRA/CRI/FIDC fora de Emp&Fin
    "outros_pass_lp":   None,
    "caixa_equiv":     105.1,    # Caixa e equiv. (Nota 5 — excl. LP de aplic.)
    "aplic_cp":       1627.1,    # Aplicações Financeiras CP (Nota 5)
    "receita_liq":    4109.4,    # LTM explícita no release 1T26 (PMR, pág.10)
    "ebit":           1023.0,    # LTM: 1002,0 − 304,6 + 325,6
    "da":              420.8,    # LTM: 427,4 − 109,4 + 102,8
    "desp_fin":        887.3,    # LTM excl. IFRS16: 858,7 − 188,9 + 217,5
    "rec_fin":         272.7,    # LTM: 252,5 − 53,8 + 74,0
    "capex":           220.8,    # LTM: 217,3 − 51,0 + 54,5

    # ── Wall of Maturities (Emp.&Fin. apenas, R$ mm) ─────────────────────────
    # Anos: 2025 | 2026 | 2027 | 2028 | 2029 | 2030..2036+
    "wall": [None, 474.8, 1625.0, 1224.8, 1174.4,
             None, None, None, None, None, None, None],

    # ── Covenants & Capex ────────────────────────────────────────────────────
    "cv_instrumento":  "Debs 4ª-8ª (Ânima); Debs 2ª (Inspirali); BB Capital de Giro; IFC",
    "cv_dl_ebitda":    "≤ 3,0x (Debs Ânima); ≤ 3,5x (Inspirali, BB-Rede)",
    "cv_outro":        "EBITDA Ajustado Pf / Desp.Fin. ≥ 1,3; IFC: Caixa/Div.CP ≥ 1,20",
    "cv_headroom":     "DL/EBITDA ex-IFRS16 LTM = 2,39x vs limite 3,0x → headroom 0,61x",
    "cv_capex_12m":    None,
    "cv_capex_24m":    None,
    "cv_ma":           "FASEH 94,42% (R$45,3mm) — concluído fev/2026",
    "cv_impacto":      "Marginal — deal pequeno e já liquidado",
    "cv_obs":          "Sem descumprimento de covenants em 31/03/2026 (Nota 12 DF 1T26)",

    # ── Cronograma de Amortização ─────────────────────────────────────────────
    # Cols: Tipo | Série | DataEmissão | Saldo | Moeda | VencFinal |
    #        PróxAmort | ValorPróxAmort | Taxa | Indexador | Obs
    "cronograma": [
        ("Debênture", "4ª Emissão — Série 1",    date(2022,12,28), 596.6, "BRL", date(2027,12,13), date(2027,12,13), 596.6, "CDI+1,65% aa", "CDI",  "Bullet. Covenant: DL/EBITDA Ajustado Pf ≤3,0x; EBITDA Pf/Desp.Fin. ≥1,3"),
        ("Debênture", "4ª Emissão — Série 2",    date(2022,12,28), 211.6, "BRL", date(2029,12,13), date(2028,12,13), None,  "IPCA+8,05% aa","IPCA", "Amort. anual: dez/2028 e dez/2029. Covenant: igual Série 1"),
        ("Debênture", "5ª Emissão",               date(2024,5,22),  211.4, "BRL", date(2029,5,15),  date(2026,5,15),  None,  "CDI+1,92% aa", "CDI",  "Amort. anual a partir mai/2026. Covenant: DL/EBITDA Ajustado Pf ≤3,0x"),
        ("Debênture", "6ª Emissão",               date(2024,8,20),  319.8, "BRL", date(2029,8,16),  date(2026,8,16),  None,  "CDI+1,92% aa", "CDI",  "Amort. semestral a partir fev/2026. Covenant: DL/EBITDA Ajustado Pf ≤3,0x"),
        ("Debênture", "7ª Emissão",               date(2025,5,21),  157.4, "BRL", date(2029,11,14), date(2028,5,14),  None,  "CDI+1,60% aa", "CDI",  "2 parcelas: mai/2028 e nov/2029. Covenant: DL/EBITDA Ajustado Pf ≤3,0x"),
        ("Debênture", "8ª Emissão",               date(2025,10,15), 315.3, "BRL", date(2029,10,15), date(2028,10,15), None,  "CDI+1,50% aa", "CDI",  "Amort. a partir out/2028. Covenant: DL/EBITDA Ajustado Pf ≤3,0x"),
        ("Bancário",  "Banco ABC 3",              date(2024,9,30),  104.4, "BRL", date(2027,9,27),  date(2026,6,27),  None,  "CDI+1,75% aa", "CDI",  "Amort. trim. desde dez/2025. Covenant: DL/EBITDA Ajustado Pf ≤3,0x"),
        ("Bancário",  "BB Capital de Giro 1",     date(2023,6,28),   87.2, "BRL", date(2029,9,25),  date(2027,3,25),  None,  "CDI+1,65% aa", "CDI",  "Amort. semestral. Covenant: DL/EBITDA ajustado <3,5x aa"),
        ("Bancário",  "BB Capital de Giro 2",     date(2023,4,27),   81.0, "BRL", date(2029,9,28),  date(2027,3,28),  None,  "CDI+1,65% aa", "CDI",  "Amort. semestral. Sem covenant explícito"),
        ("Bancário",  "BB Capital de Giro 3",     date(2023,7,4),    88.3, "BRL", date(2029,9,25),  date(2027,3,25),  None,  "CDI+1,65% aa", "CDI",  "Amort. semestral. Covenant: DL/EBITDA ajustado <3,5x aa"),
        ("Empréstimo","IFC (USD)",                date(2020,7,10),   81.1, "USD", date(2028,3,15),  None,             None,  "SOFR+2,43% aa","SOFR", "Moeda original USD; convertido BRL pelo câmbio do balanço. Covenant: Caixa/Div.CP ≥1,20; DL/EBITDA ajustado <3; EBITDA ajustado/Desp.Fin.líquida ≥1,30"),
        ("Debênture", "Inspirali — 2ª Emissão",   date(2024,5,27), 2110.6, "BRL", date(2029,5,15),  date(2027,5,15),  None,  "CDI+1,65% aa", "CDI",  "Amort. anual mai/2027→2029. Covenant: DL/EBITDA Ajustado Pf ≤3,5x; EBITDA Pf/Desp.Fin. ≥1,3"),
        ("Bancário",  "BB Capital de Giro (Rede)",date(2021,6,25),  124.3, "BRL", date(2028,9,25),  date(2026,9,25),  None,  "CDI+1,65% aa", "CDI",  "Amort. semestral desde mar/2025. Covenant: DL/EBITDA <3,5x"),
    ],

    # ── Disclaimers ───────────────────────────────────────────────────────────
    "disclaimers": [
        ("Ânima","Período / LTM","Calculado","LTM = FY2025 (4T25 release) − 1T25 + 1T26 (ITR 1T26)","Anima_Release_4T25_PT.pdf pg.1; Anima_Release_1T26.pdf pg.1","12/03/2026; 06/05/2026"),
        ("Ânima","Receita Líquida LTM","Calculado","R$4.109,4mm — figura explícita do release 1T26 (PMR pág.10). Cálculo via fórmula: 4.023,7−1.040,1+1.120,4=4.104,0mm (dif. de R$5,4mm por arredondamento nos trimestrais).","Anima_Release_1T26.pdf pg.10","06/05/2026"),
        ("Ânima","EBIT LTM","Calculado","1.002,0−304,6+325,6=R$1.023,0mm. EBIT = Resultado antes do resultado financeiro (IFRS). Inclui equivalência patrimonial.","Anima_Release_4T25_PT.pdf pg.20; Anima_DF_1T26.pdf pg.5","12/03/2026; 06/05/2026"),
        ("Ânima","D&A LTM","Calculado","427,4−109,4+102,8=R$420,8mm. D&A inclui amortização de ativo de direito de uso (IFRS16). Hierarquia (3): calculado.","Anima_Release_4T25_PT.pdf pg.15; Anima_Release_1T26.pdf pg.9","12/03/2026; 06/05/2026"),
        ("Ânima","EBITDA LTM (col V)","Calculado","EBITDA (fórmula) = EBIT+D&A = R$1.443,8mm. Release: EBITDA Ajustado ex-IFRS16 LTM = R$1.222,2mm (excl. IFRS16 e ajustes). Hierarquia usada: (3) EBIT+D&A.","Anima_Release_1T26.pdf pg.10","06/05/2026"),
        ("Ânima","Desp. Financ. Bruta LTM","Calculado","Excluída despesa IFRS16 de arrendamento: (1.001,6−142,9)−(225,0−36,1)+(259,4−41,9)=R$887,3mm. Linha 'Desp. comissões/juros empréstimos' pode incluir efeito cambial do swap IFC (não separável).","Anima_Release_4T25_PT.pdf pg.14; Anima_DF_1T26.pdf pg.8","12/03/2026; 06/05/2026"),
        ("Ânima","Rec. Financeira LTM","Calculado","252,5−53,8+74,0=R$272,7mm.","Anima_Release_4T25_PT.pdf pg.14; Anima_Release_1T26.pdf pg.8","12/03/2026; 06/05/2026"),
        ("Ânima","Capex LTM","Calculado","217,3−51,0+54,5=R$220,8mm. Capex total (imobilizado+intangível), excluído M&A (FASEH R$45,3mm e outras aquisições).","Anima_Release_4T25_PT.pdf pg.18; Anima_Release_1T26.pdf pg.12","12/03/2026; 06/05/2026"),
        ("Ânima","Caixa e Aplicações CP","Ajustado","Aplicações LP (R$25,7mm) excluídas do caixa CP. R$31,8mm das aplicações estão em garantia (parcialmente restritas) — mantidas por representar <2% do total.","Anima_DF_1T26.pdf Nota 5 pg.13","06/05/2026"),
        ("Ânima","Emp. & Fin. CP/LP","Ajustado","Excluídos derivativos CP (R$20,1mm) e LP (R$21,8mm) conforme R10. Inclui debêntures e empréstimos bancários consolidados (Nota 12).","Anima_DF_1T26.pdf Nota 12 pg.26; Balanço pg.4","06/05/2026"),
        ("Ânima","Valor Próx. Amortização","Não disponível","Valores exatos por parcela não detalhados no Nota 12 do ITR 1T26 para instrumentos com amortização parcial. Deixados em branco.","Anima_DF_1T26.pdf Nota 12 pg.26-27","06/05/2026"),
        ("Ânima","Guidance Capex 12/24m","Não disponível","Empresa não divulga guidance de capex explícito. Histórico: ~5,4% da receita líquida (2025).","Anima_Release_4T25_PT.pdf pg.18","12/03/2026"),
    ],

    # ── Oportunidades de Deal ─────────────────────────────────────────────────
    "oportunidades": [
        ("Refinanciamento Preventivo","Debêntures CDI+ 5-7 anos","R$ 1.600–1.800mm",
         "Wall de vencimentos em 2027 representa R$1.625mm (36% da dívida total): Debs 4ª Série 1 (R$597mm bullet dez/2027) + Banco ABC (R$104mm set/2027) + parcelas BB. DL/EBITDA de 2,39x oferece headroom confortável para nova emissão a custo competitivo.",
         "Debêntures CDI+ em 5-7 anos com amortização anual a partir do ano 3. Possível estrutura IPCA+ para tranche de prazo mais longo (7+ anos) se condições de mercado permitirem.",
         "Alta","2H26 — emissão preferencial antes de set/2026 para evitar sazonalidade negativa do 2T"),
        ("Emissão Incremental / M&A","Debêntures Inspirali CDI+ ou IPCA+","R$ 500–800mm",
         "Inspirali (medicina) tem perfil de crédito superior: margem operacional de 52,8% e crescimento de receita de 9,8% em 2025. Emissão segregada captura prêmio de crédito mais favorável vs. holding. Potencial de elegibilidade como infraestrutura educacional médica (Lei 12.431) para cursos de medicina.",
         "Debêntures Inspirali com covenant específico ≤3,5x DL/EBITDA; possível estruturação como IPCA+ via Lei 12.431 caso aprovação regulatória. Garantia da holding Ânima como credit support.",
         "Média","1H27 — após estabilização regulatória do Novo Marco Regulatório do EAD"),
    ],
}

# ══════════════════════════════════════════════════════════════════════════════
#  FUNCTIONS — Fill each sheet
# ══════════════════════════════════════════════════════════════════════════════

def fill_inputs_dfs(ws, data):
    row = find_row(ws, data["name"])
    if not row:
        print(f"  ⚠ '{data['name']}' não encontrado em Inputs—DFs")
        return
    # Cols B..P = 2..16
    values = [
        data["periodo"],        # B
        data["emp_fin_cp"],     # C
        data["emp_fin_lp"],     # D
        data["arrendam_cp"],    # E
        data["arrendam_lp"],    # F
        data["outros_pass_cp"], # G
        data["outros_pass_lp"], # H
        data["caixa_equiv"],    # I
        data["aplic_cp"],       # J
        data["receita_liq"],    # K
        data["ebit"],           # L
        data["da"],             # M
        data["desp_fin"],       # N
        data["rec_fin"],        # O
        data["capex"],          # P
    ]
    for col_idx, val in enumerate(values, start=2):
        inp(ws, row, col_idx, val)
    print(f"  ✓ Inputs—DFs: '{data['name']}' linha {row}")


def fill_wall_maturities(ws, data):
    row = find_row(ws, data["name"])
    if not row:
        print(f"  ⚠ '{data['name']}' não encontrado em Wall of Maturities")
        return
    # Cols C..N = 3..14 (anos 2025..2036+)
    for col_idx, val in enumerate(data["wall"], start=3):
        if val is not None:
            inp(ws, row, col_idx, val)
    print(f"  ✓ Wall of Maturities: '{data['name']}' linha {row}")


def fill_covenants(ws, data):
    row = find_row(ws, data["name"])
    if not row:
        print(f"  ⚠ '{data['name']}' não encontrado em Covenants & Capex")
        return
    values = [
        data["cv_instrumento"],  # C
        data["cv_dl_ebitda"],    # D
        data["cv_outro"],        # E
        data["cv_headroom"],     # F
        data["cv_capex_12m"],    # G
        data["cv_capex_24m"],    # H
        data["cv_ma"],           # I
        data["cv_impacto"],      # J
        data["cv_obs"],          # K
    ]
    for col_idx, val in enumerate(values, start=3):
        if val is not None:
            c = ws.cell(row=row, column=col_idx, value=val)
            c.alignment = WRAP
    print(f"  ✓ Covenants & Capex: '{data['name']}' linha {row}")


def fill_cronograma(ws, data):
    """Insert rows for company instruments and fill them."""
    anchor = find_row(ws, data["name"])
    if not anchor:
        print(f"  ⚠ '{data['name']}' não encontrado em Cronograma")
        return

    instruments = data["cronograma"]
    n = len(instruments)

    # Find how many blank rows already exist after anchor
    blanks = 0
    for r in range(anchor + 1, anchor + 20):
        if ws.cell(row=r, column=1).value:
            break
        blanks += 1

    # Insert rows if needed (keep at least 1 row for anchor itself)
    extra_needed = n - 1  # anchor row handles first instrument
    if extra_needed > blanks:
        ws.insert_rows(anchor + 1, extra_needed - blanks)

    for i, inst in enumerate(instruments):
        row = anchor + i
        tipo, serie, data_em, saldo, moeda, venc, prox_amort, valor_prox, taxa, indexador, obs = inst
        ws.cell(row=row, column=1, value=data["name"])
        ws.cell(row=row, column=2, value=data["sector"])
        ws.cell(row=row, column=3, value=tipo)
        ws.cell(row=row, column=4, value=serie)
        ws.cell(row=row, column=5, value=data_em)
        ws.cell(row=row, column=6, value=saldo)
        ws.cell(row=row, column=7, value=moeda)
        ws.cell(row=row, column=8, value=venc)
        if prox_amort:
            ws.cell(row=row, column=9,  value=prox_amort)
        if valor_prox:
            ws.cell(row=row, column=10, value=valor_prox)
        ws.cell(row=row, column=11, value=taxa)
        ws.cell(row=row, column=12, value=indexador)
        c = ws.cell(row=row, column=13, value=obs)
        c.alignment = WRAP

    print(f"  ✓ Cronograma: '{data['name']}' — {n} instrumentos (linhas {anchor}–{anchor+n-1})")


def fill_disclaimers(ws, data, start_row):
    """Append disclaimers for a company starting at start_row."""
    headers = ["Empresa","Campo Afetado","Situação","Detalhamento","Fonte (doc + página)","Data do Documento"]
    if start_row == 2:  # Write headers on first call
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 55
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 18

    row = start_row
    for disc in data["disclaimers"]:
        for col, val in enumerate(disc, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = WRAP
        row += 1
    return row


def fill_oportunidades(ws, data, start_row):
    """Append deal opportunities for a company."""
    headers = ["Empresa","Setor","Tipo de Oportunidade","Instrumento Sugerido",
               "Tamanho Estimado","Racional","Estrutura Sugerida","Prioridade","Timing"]
    if start_row == 2:
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER
        widths = [12,12,24,28,18,60,45,10,35]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    row = start_row
    for op in data["oportunidades"]:
        tipo, instr, tam, racional, estrut, prio, timing = op
        vals = [data["name"], data["sector"], tipo, instr, tam, racional, estrut, prio, timing]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = WRAP
        row += 1
    return row


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN — Process companies
# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
#  DATA — YDUQS (1T26 LTM)
# ══════════════════════════════════════════════════════════════════════════════
YDUQS = {
    "name":   "Yduqs",
    "sector": "Educação",

    # ── Inputs — DFs (R$ mm) ─────────────────────────────────────────────────
    "periodo":        "1T26 LTM",
    "emp_fin_cp":      582.3,    # Emp & Fin (excl. IFRS16 arrendamento CP)
    "emp_fin_lp":     3184.4,    # Emp & Fin (excl. IFRS16 arrendamento LP)
    "arrendam_cp":     264.6,    # IFRS16 CP (Financiamento por Arrendamento)
    "arrendam_lp":    1338.3,    # IFRS16 LP
    "outros_pass_cp":   None,    # M&A payables CP <R$50mm — omitido
    "outros_pass_lp":   None,    # M&A payables LP e Opções: registrar no Disclaimer
    "caixa_equiv":     571.4,    # Caixa e Equivalentes consolidado
    "aplic_cp":        509.8,    # Aplicações Financeiras CP (Fundo de Investimento)
    "receita_liq":    5543.4,    # LTM: 5.521,7 − 1.487,1 + 1.508,8
    "ebit":            822.6,    # LTM: 887,3 − 295,8 + 231,1
    "da":              804.8,    # LTM: 815,6 − 207,4 + 196,6
    "desp_fin":        937.0,    # LTM excl. IFRS16: (1.139,5−174,8)−(300,9−43,5)+(274,2−44,5)
    "rec_fin":         372.6,    # LTM: 391,7 − 113,1 + 94,0
    "capex":           460.3,    # LTM: 461,0 − 114,3 + 113,6

    # ── Wall of Maturities (principal schedule, release pág.31) ──────────────
    # Anos: 2025 | 2026 | 2027 | 2028 | 2029 | 2030 | 2031 | 2032..2036+
    "wall": [None, 200.0, 457.0, 988.0, 1103.0, 703.0, 150.0, None, None, None, None, None],

    # ── Covenants & Capex ────────────────────────────────────────────────────
    "cv_instrumento":  "Debêntures VII-XII; CCB Itaú",
    "cv_dl_ebitda":    "Não disponível sem Notas Explicativas — DL/EBITDA ajustado LTM = 1,53x",
    "cv_outro":        "Não disponível sem Notas Explicativas",
    "cv_headroom":     "DL/EBITDA Ajustado (ex-IFRS16) LTM = 1,53x — leverage muito baixa",
    "cv_capex_12m":    None,
    "cv_capex_24m":    None,
    "cv_ma":           "Unifametro (medicina, IDOMED) — adquirida 1T26; pagamento parcial R$75,5mm",
    "cv_impacto":      "Marginal — leverage aumentou de 1,59x para 1,53x (redução no período)",
    "cv_obs":          "FCA guidance 2026: R$520-620mm. Custo médio dívida: CDI+0,99%. Sem info covenant nos documentos disponíveis.",

    # ── Cronograma de Amortização ─────────────────────────────────────────────
    # Saldos aprox. (pro-rateado do 'saldo a pagar' do release vs book value balanço)
    # Vencimentos estimados por prazo médio a partir de 31/03/2026
    # Nota: saldos = estimativas; datas exatas requerem Nota Explicativa
    "cronograma": [
        ("Bancário",  "CCB - Itaú",              None, 208.9, "BRL", date(2026,5,31),  date(2026,5,31),  208.9, "CDI+1,15% aa", "CDI",  "Prazo médio 0,2 anos (CP). Saldo e data: estimados por prazo médio e pro-rata"),
        ("Debênture", "7ª Emissão (Única)",       None, 312.7, "BRL", date(2027,12,31), None,             None,  "CDI+0,78% aa", "CDI",  "Prazo médio 1,7 anos. Data e saldo: estimados"),
        ("Debênture", "8ª Emissão (Única)",       None, 498.7, "BRL", date(2030,6,30),  None,             None,  "CDI+0,85% aa", "CDI",  "Prazo médio 4,2 anos. Data e saldo: estimados"),
        ("Debênture", "9ª Emissão — 1ª série",   None, 297.9, "BRL", date(2028,9,30),  None,             None,  "CDI+0,82% aa", "CDI",  "Prazo médio 2,5 anos. Data e saldo: estimados"),
        ("Debênture", "9ª Emissão — 2ª série",   None, 333.9, "BRL", date(2028,3,31),  None,             None,  "CDI+0,90% aa", "CDI",  "Prazo médio 2,0 anos. Data e saldo: estimados"),
        ("Debênture", "9ª Emissão — 3ª série",   None, 112.0, "BRL", date(2030,3,31),  None,             None,  "CDI+0,98% aa", "CDI",  "Prazo médio 4,0 anos. Data e saldo: estimados"),
        ("Debênture", "10ª Emissão (Única)",      None,1173.1, "BRL", date(2028,9,30),  None,             None,  "CDI+1,25% aa", "CDI",  "Prazo médio 2,5 anos. Maior instrumento (31% da dívida). Data e saldo: estimados"),
        ("Debênture", "11ª Emissão (Única)",      None, 313.1, "BRL", date(2031,5,31),  None,             None,  "CDI+1,05% aa", "CDI",  "Prazo médio 5,2 anos. Data e saldo: estimados"),
        ("Debênture", "12ª Emissão (Única)",      None, 516.4, "BRL", date(2030,1,31),  None,             None,  "CDI+0,70% aa", "CDI",  "Prazo médio 3,7 anos. Data e saldo: estimados"),
    ],

    # ── Disclaimers ───────────────────────────────────────────────────────────
    "disclaimers": [
        ("Yduqs","Período / LTM","Calculado","LTM = FY2025 (DFP 31/12/2025) − 1T25 + 1T26 (ITR 31/03/2026)","Complete_com_o_Docusign_Yduqs_Part_4T25_vf_C.pdf pg.19; Yduqs_ITR_1T26.pdf pg.13","DFP:31/12/2025; ITR:31/03/2026"),
        ("Yduqs","Receita Líquida LTM","Calculado","5.521,7 − 1.487,1 + 1.508,8 = R$5.543,4mm. Fonte 1T26/1T25: ITR DRE consolidada pg.13. Fonte FY2025: DFP DRE consolidada pg.19.","Yduqs_ITR_1T26.pdf pg.13; DFP 4T25 pg.19","31/03/2026; 31/12/2025"),
        ("Yduqs","EBIT LTM","Calculado","887,3 − 295,8 + 231,1 = R$822,6mm. EBIT = Resultado antes do Resultado Financeiro (IFRS).","Yduqs_ITR_1T26.pdf pg.13; DFP 4T25 pg.19","31/03/2026; 31/12/2025"),
        ("Yduqs","D&A LTM","Calculado","815,6 − 207,4 + 196,6 = R$804,8mm. Inclui amortização de ROU (IFRS16) e mais-valia/ágio. FY2025 via cash flow DFP. Hierarquia: (3) calculado.","Yduqs_ITR_1T26.pdf pg.15; DFP 4T25 pg.21","31/03/2026; 31/12/2025"),
        ("Yduqs","EBITDA LTM (col V)","Calculado","EBITDA (fórmula) = EBIT+D&A = R$1.627,4mm. Release: EBITDA Ajustado LTM = R$1.864,6mm (ajustado e com IFRS16); EBITDA Ajustado ex-IFRS16 LTM = R$1.864,6mm (ex-IFRS16). Discrepância: fórmula inclui não-recorrentes negativos de R$76,4mm no 1T26 (venda imóvel contábil). Hierarquia: (3).","Yduqs_Release_1T26.pdf pg.20,30","07/05/2026"),
        ("Yduqs","Desp. Financ. Bruta LTM","Calculado","Excluída despesa IFRS16 arrendamento. LTM = (1.139,5−174,8)−(300,9−43,5)+(274,2−44,5) = R$937,0mm. ITR usa DRE IFRS (difere do release, que exclui efeito swap e outras reclassificações).","Yduqs_ITR_1T26.pdf pg.13,15; DFP 4T25 pg.19,21","31/03/2026; 31/12/2025"),
        ("Yduqs","Rec. Financeira LTM","Calculado","391,7 − 113,1 + 94,0 = R$372,6mm (via DRE IFRS). Inclui multas/juros recebidos, aplicações e atualização monetária. FY2025 alto (R$391,7mm) reflete SELIC elevada e caixa médio ~R$1,5bi.","Yduqs_ITR_1T26.pdf pg.13; DFP 4T25 pg.19","31/03/2026; 31/12/2025"),
        ("Yduqs","Capex LTM","Calculado","461,0 − 114,3 + 113,6 = R$460,3mm. FY2025 Capex via release pág.29. 1T25/1T26 via release pág.26. Excluído M&A (Unifametro R$75,5mm).","Yduqs_Release_1T26.pdf pg.26,29","07/05/2026"),
        ("Yduqs","Outros Passivos Fin. LP","Não disponível","M&A payables LP (R$112,9mm = Preço de Aquisição a Pagar LP) e Passivo Financeiro-Opções LP (R$7,9mm) não mapeados em Outros Passivos Fin. — não são CRA/CRI/FIDC. Release inclui 'Compromissos a pagar M&A' (R$162,5mm) na Dívida Bruta mas omitidos da planilha por falta de coluna adequada.","Yduqs_ITR_1T26.pdf Balanço pg.11-12; Yduqs_Release_1T26.pdf pg.30","31/03/2026; 07/05/2026"),
        ("Yduqs","Cronograma — Saldos/Datas","Estimado","Saldos e datas de vencimento estimados por pro-rata e prazo médio divulgados no release pág.30. Saldos exatos por instrumento e datas finais requerem Nota Explicativa 12 do ITR. Datas emissão não disponíveis no release.","Yduqs_Release_1T26.pdf pg.30","07/05/2026"),
        ("Yduqs","Covenants","Não disponível","Limites de covenants financeiros não identificados nos documentos disponíveis (release e ITR). Requer leitura da Nota Explicativa de Dívida no ITR 1T26.","Yduqs_ITR_1T26.pdf Nota 12 (pg.61+)","31/03/2026"),
    ],

    # ── Oportunidades de Deal ─────────────────────────────────────────────────
    "oportunidades": [
        ("Pré-Refinanciamento","Debêntures CDI+ 5-7 anos","R$ 988–1.200mm",
         "Wall concentrada em 2028 (R$988mm de Debs IX séries 1 e 2 + parcela Debs X): R$1.500mm+ vencendo em 2028. Aproveitamento de alavancagem muito baixa (1,53x) e FCO robusto (R$457,9mm no 1T26) para alongar perfil a custo favorável. Guidance FCA de R$520-620mm em 2026 sinaliza forte geração de caixa estrutural.",
         "Debêntures CDI+ em 5-7 anos com amortização anual a partir do ano 3. Possível estrutura em duas tranches: Série 1 (CDI+) para varejo e Série 2 (IPCA+) para investidores institucionais de prazo longo.",
         "Alta","2H26 — antes de eventual estresse de mercado e aproveitando janela pós-ciclo de alta Selic"),
        ("Crescimento Premium","Debêntures IPCA+ (Lei 12.431)","R$ 500–700mm",
         "IDOMED (medicina) crescendo 10% em ROL e 8% em EBITDA. Ibmec crescendo 23% em ROL e 34% em EBITDA. Expansão em campi médicos/negócios de alta qualidade pode ser financiada por debêntures de infraestrutura se classificada como infraestrutura educacional (medicina). Baixa alavancagem permite incremento de ~R$700mm sem violar covenants típicos de 3,0x.",
         "Debêntures Lei 12.431 (IPCA+ isento IR para PF) se elegíveis; ou CDI+ em SFN com garantia fiduciária das marcas/receitas Ibmec/IDOMED. Avaliação de elegibilidade regulatória necessária.",
         "Média","2026-2027 — pós aprovação de novos campi e/ou M&A adicional"),
    ],
}

# ══════════════════════════════════════════════════════════════════════════════
#  DATA — AFYA (1Q26 LTM)
# ══════════════════════════════════════════════════════════════════════════════
AFYA = {
    "name":   "Afya",
    "sector": "Educação",

    # ── Inputs — DFs (R$ mm, from 1Q26 Press Release financial tables) ────────
    "periodo":        "1T26 LTM",
    "emp_fin_cp":      132.1,    # Loans and financing CP (Notas Comerciais + IFC)
    "emp_fin_lp":     1992.4,    # Loans and financing LP
    "arrendam_cp":      55.5,    # Lease liabilities CP (IFRS16)
    "arrendam_lp":    1021.6,    # Lease liabilities LP (IFRS16)
    "outros_pass_cp":   57.3,    # Accounts payable to selling shareholders CP (M&A earnouts)
    "outros_pass_lp":  302.3,    # Accounts payable to selling shareholders LP
    "caixa_equiv":    1332.9,    # Cash and cash equivalents (incl. cash equivalents R$1.307mm)
    "aplic_cp":         None,    # Included in Cash and Cash Equivalents
    "receita_liq":    3773.6,    # LTM: 3.697,3 − 936,4 + 1.012,7
    "ebit":           1234.2,    # LTM: 1.213,1 − 372,5 + 393,6 (Operating income IFRS)
    "da":              374.6,    # LTM: 373,3 − 91,8 + 93,1
    "desp_fin":         None,    # FY2025 Desp.Fin. bruta não disponível separada → Disclaimer
    "rec_fin":          None,    # Idem
    "capex":           293.0,    # LTM: 304,4 − 56,2 + 44,8 (excl. licença M&A FUNIC R$99,6mm)

    # ── Wall of Maturities (Emp.&Fin. apenas, R$ mm) ─────────────────────────
    # Notas Comerciais 1ª (R$498mm bullet out/2028) + 2ª (R$996mm bullet out/2030)
    # IFC (R$530mm, amortização trimestral, ~2028-2029)
    "wall": [None, 132.1, 132.0, 630.0, 134.0, 996.0, None, None, None, None, None, None],

    # ── Covenants & Capex ────────────────────────────────────────────────────
    "cv_instrumento":  "Notas Comerciais 1ª e 2ª Série; IFC; Contas a pagar a vendedores",
    "cv_dl_ebitda":    "Net Debt / Adjusted EBITDA ≤ 3,0x (excl. IFRS16, Notas Comerciais)",
    "cv_outro":        "Conforme Commercial Notes (press release 4T25, pg.2)",
    "cv_headroom":     "DL/EBITDA Ajustado mid-guidance 2026 = 0,7x vs limite 3,0x → headroom 2,3x",
    "cv_capex_12m":    360.0,    # Midpoint guidance 2026: R$340-380mm
    "cv_capex_24m":    None,
    "cv_ma":           "FUNIC (60 vagas medicina, dez/2025, licença R$99,6mm); 63 vagas MEC Abaetetuba fev/2026",
    "cv_impacto":      "Impacto marginal — alavancagem reduzindo de 0,8x para 0,7x",
    "cv_obs":          "Rating: AAA.br (Moody's, mai/2026). DL/EBITDA ex-IFRS16 1Q26 = 0,7x. FY2025 compliant.",

    # ── Cronograma de Amortização ─────────────────────────────────────────────
    "cronograma": [
        ("Debênture","Notas Comerciais 1ª Série", date(2025,10,15), 498.0,"BRL",date(2028,10,15),date(2028,10,15),498.0,"CDI+0,70% aa","CDI",  "Bullet. Emissão via Opea Securitizadora. Covenant: DL (ex-IFRS16)/Adj.EBITDA ≤3,0x anual"),
        ("Debênture","Notas Comerciais 2ª Série", date(2025,10,15), 996.0,"BRL",date(2030,10,15),date(2030,10,15),996.0,"CDI+0,85% aa","CDI",  "Bullet. Mesmo covenant da 1ª Série. Total emissão R$1,5bi (R$500mm+R$1.000mm)"),
        ("Empréstimo","IFC",                      date(2020,1,1),   530.0,"USD",date(2029,1,31), None,            None, "SOFR+spread aa","SOFR", "Moeda original USD; convertido BRL pelo câmbio do balanço. Amortização trimestral. Duration 2,8 anos (mar/2026). Data emissão aproximada."),
        ("Outro",    "Contas a Pagar a Vendedores (M&A)",None,       360.0,"BRL",date(2030,6,30), None,            None, "100% CDI aa",   "CDI",  "Earn-outs e parcelas devidas a ex-sócios de empresas adquiridas. Incluído na Dívida Bruta do release. Duration 4,2 anos. Não é dívida financeira clássica."),
    ],

    # ── Disclaimers ───────────────────────────────────────────────────────────
    "disclaimers": [
        ("Afya","Período / LTM","Calculado","LTM = FY2025 (4Q25 press release 12/03/2026) − 1Q25 + 1Q26 (press release 07/05/2026). Valores em R$ mil, convertidos para mm.","Afya_Release_1T26.pdf; Afya Limited...FY2025 Results.pdf","07/05/2026; 12/03/2026"),
        ("Afya","Receita Líquida LTM","Calculado","3.697,3 − 936,4 + 1.012,7 = R$3.773,6mm. Receita total reportada (IFRS), excl. intercompany.","Afya_Release_1T26.pdf Table 6; FY2025 Results Table 6","07/05/2026; 12/03/2026"),
        ("Afya","EBIT LTM","Calculado","1.213,1 − 372,5 + 393,6 = R$1.234,2mm. EBIT = 'Operating income' do DRE IFRS. FY2025 derivado: Net income + Taxes + Net fin result − Share of associate = R$1.213,1mm.","Afya_Release_1T26.pdf pg.13; FY2025 Table 7","07/05/2026; 12/03/2026"),
        ("Afya","D&A LTM","Calculado","373,3 − 91,8 + 93,1 = R$374,6mm. Inclui amortização de ROU (IFRS16). Hierarquia: (3) calculado via EBITDA reconciliation Table 7.","Afya_Release_1T26.pdf pg.7; FY2025 Table 7","07/05/2026; 12/03/2026"),
        ("Afya","Desp. Financ. Bruta / Rec. Fin. LTM","Não disponível","Receitas e Despesas Financeiras FY2025 não disponíveis de forma separada no press release anual (apenas resultado financeiro líquido = R$366,1mm é divulgado). Requer Form 20-F (SEC) ou demonstração de resultado completa. Campos deixados em branco.","Afya Limited...FY2025 Results.pdf Table 7","12/03/2026"),
        ("Afya","Capex LTM","Calculado","304,4 − 56,2 + 44,8 = R$293,0mm. FY2025 Capex total = R$404mm; excluída licença FUNIC (R$99,6mm = M&A per Rule Capex). 1Q25/1Q26 via Table 12 do press release.","Afya_Release_1T26.pdf Table 12; FY2025 Results Table 12","07/05/2026; 12/03/2026"),
        ("Afya","Caixa e Aplicações CP","Nota","Cash and Cash Equivalents inclui Cash Equivalents = R$1.307mm (fundos de alta liquidez). Não há linha separada de 'Aplicações Financeiras CP' no balanço do press release.","Afya_Release_1T26.pdf Table 11 pg.12","07/05/2026"),
        ("Afya","Outros Passivos Fin.","Ajustado","'Accounts payable to selling shareholders' (total R$359,7mm) incluído em Outros Passivos Fin. conforme a própria empresa inclui na Dívida Bruta (Table 10). Não é CRA/CRI/FIDC mas é passivo financeiro oneroso (CDI 100%).","Afya_Release_1T26.pdf Table 10-11 pg.8,12","07/05/2026"),
        ("Afya","Wall of Maturities","Estimado","Amortização IFC estimada com base em duration 2,8 anos e balance R$530mm (parcelas anuais ~R$132mm). Datas exatas requerem Nota Explicativa do DF 1Q26.","Afya_Release_1T26.pdf Table 10","07/05/2026"),
        ("Afya","Cronograma — IFC","Estimado","Data de emissão IFC e datas de amortização: estimadas. Moeda original USD (convertida BRL pelo câmbio do balanço). SOFR+spread: custo total = 15,8% aa em 1Q26 (CDI+108%).","Afya_Release_1T26.pdf Table 10","07/05/2026"),
    ],

    # ── Oportunidades de Deal ─────────────────────────────────────────────────
    "oportunidades": [
        ("Refinanciamento IFC / Diversificação","Debêntures CDI+ ou IPCA+ 5-7 anos","R$ 500–700mm",
         "IFC (R$530mm) com amortização trimestral e custo 108% do CDI (~15,8% aa) é o instrumento mais caro da estrutura. Refinanciamento via debêntures locais reduziria spread e alongaria perfil. DL/EBITDA de 0,7x (headroom excepcional vs covenant 3,0x) permite emissão incremental se necessário.",
         "Debêntures CDI+ em 5-6 anos com amortização anual a partir do ano 3 (similar às Notas Comerciais existentes). Possível estrutura IPCA+ se Afya qualificar como infraestrutura educacional médica (3.768 vagas de medicina = argumento para Lei 12.431).",
         "Média","2026-2027 — quando IFC amortizar parcialmente e janela de mercado permitir"),
        ("Expansão M&A","Debêntures CDI+ Incrementais","R$ 800–1.200mm",
         "Afya executa consistente ciclo de M&A (FUNIC 2025, UNIDOM 2024, FIP Guanambi 2024). Com FCF de R$1.056mm em 2025 e guidance 2026 de R$1.700-1.800mm EBITDA, há espaço para leverage adicional de até 2x (DL de ~R$3.400mm) vs atual R$1.151mm. Nova captação financiaria aquisições de vagas médicas a prêmio regulatório.",
         "Debêntures incrementais CDI+ para aquisição de escolas de medicina com vagas MEC (escassas e com deságio natural para compradores). Estrutura de holding (Afya Brazil) com garantias das subsidiárias operacionais.",
         "Alta","2H26-1H27 — após conclusão do ciclo de integração FUNIC e aprovação de novas vagas MEC"),
    ],
}

COMPANIES = [ANIMA, YDUQS, AFYA]  # Add more companies here as data is extracted

disc_row  = 2
deal_row  = 2

for co in COMPANIES:
    print(f"\n▶ Processando: {co['name']}")
    fill_inputs_dfs(ws_inputs, co)
    fill_wall_maturities(ws_wall, co)
    fill_covenants(ws_cov, co)
    fill_cronograma(ws_crono, co)
    disc_row = fill_disclaimers(ws_disc, co, disc_row)
    deal_row = fill_oportunidades(ws_deal, co, deal_row)

# Adjust column widths for key sheets
for col in range(1, 17):
    ws_inputs.column_dimensions[get_column_letter(col)].width = 14
ws_inputs.column_dimensions['A'].width = 20

wb.save(OUTPUT)
print(f"\n✅ Excel salvo em:\n   {OUTPUT}")
print(f"   Empresas processadas: {len(COMPANIES)}")
